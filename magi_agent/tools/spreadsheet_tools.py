from __future__ import annotations

import csv
import hashlib
import io
import json
import posixpath
import re
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path

from .context import ToolContext
from .result import ToolResult


_DEFAULT_MAX_ROWS = 100
_DEFAULT_MAX_COLS = 20
_MAX_ROWS = 10_000
_MAX_COLS = 200
_MAX_CELLS = 50_000
_MAX_BYTES = 1024 * 1024
_MAX_PREVIEW_CELL_CHARS = 240
_MAX_PREVIEW_MARKDOWN_BYTES = 4096
_DIGEST_PREFIX_LENGTH = 24
_SENSITIVE_TEXT_RE = re.compile(
    r"(?:"
    r"authorization\s*:\s*[^\n\r]+|"
    r"\bbearer\s+[A-Za-z0-9._~+/=-]+|"
    r"\bcookie\s*:\s*[^\n\r]+|"
    r"\bset-cookie\s*:\s*[^\n\r]+|"
    r"\bsid=[A-Za-z0-9._-]+|"
    r"\bsk-[A-Za-z0-9._-]+|"
    r"gh[opusr]_[A-Za-z0-9_]+|"
    r"github_pat_[A-Za-z0-9_]+|"
    r"xox[a-z]-[A-Za-z0-9._-]+|"
    r"AKIA[0-9A-Z]{8,}|"
    r"AIza[A-Za-z0-9_-]+|"
    r"(?:(?:api[_-]?key|auth(?:orization)?|cookie|password|secret|token|"
    r"session(?:[_-]?(?:key|id)|key|id))"
    r"\s*[:=]\s*|session\s*=\s*)[^\s,;}\"']+|"
    r"/Users(?:/[^\s,;}\"']*)?|"
    r"/home(?:/[^\s,;}\"']*)?|"
    r"/workspace(?:/[^\s,;}\"']*)?|"
    r"/data/bots(?:/[^\s,;}\"']*)?|"
    r"/private/var(?:/[^\s,;}\"']*)?|"
    r"/var/lib/kubelet(?:/[^\s,;}\"']*)?|"
    r"raw[_ -]?(?:tool|child|prompt|transcript|output|result|log|args|text)|"
    r"hidden[_ -]?reasoning|chain[_ -]?of[_ -]?thought"
    r")",
    re.IGNORECASE,
)
_SENSITIVE_PATH_PART_RE = re.compile(
    r"(?:"
    r"^\.|"
    r"(?:^|[._/-])(?:auth|config|cookie|credential|credentials|env|keys?|kube|"
    r"kubeconfig|password|private(?:key)?|secrets?|sessions?|tokens?|api[_-]?keys?)"
    r"(?:[._/-]|s?(?:\\.[A-Za-z0-9]+)?$)|"
    r"^(?:id_rsa|id_dsa|id_ecdsa|id_ed25519|\\.netrc|\\.npmrc|\\.pypirc)$"
    r")",
    re.IGNORECASE,
)
_SENSITIVE_TABLE_KEY_FRAGMENTS = frozenset(
    {
        "apikey",
        "auth",
        "authorization",
        "cookie",
        "credential",
        "credentials",
        "password",
        "privatekey",
        "secret",
        "servicekey",
        "session",
        "sessionid",
        "sessionkey",
        "sid",
        "token",
    }
)


@dataclass(frozen=True)
class _ResolvedPath:
    path: Path
    relative: str
    path_ref: str


class _SpreadsheetPolicyError(ValueError):
    def __init__(self, reason_code: str) -> None:
        super().__init__(reason_code)
        self.reason_code = reason_code


class _SpreadsheetDataError(ValueError):
    def __init__(self, reason_code: str) -> None:
        super().__init__(reason_code)
        self.reason_code = reason_code


def csv_read(arguments: Mapping[str, object], context: ToolContext) -> ToolResult:
    tool_name = "csv_read"
    path_text = _string_arg(arguments, "path")
    if path_text is None:
        return _blocked_result(tool_name, "path_required")
    try:
        root = _workspace_root(context)
        resolved = _resolve_workspace_path(root, path_text, must_exist=True)
        extension_error = _extension_error(resolved.relative)
        if extension_error is not None:
            return _blocked_result(tool_name, extension_error, _extension_message(extension_error))
        raw = _read_bounded_bytes(resolved.path)
    except _SpreadsheetPolicyError as error:
        return _blocked_result(tool_name, error.reason_code)
    except OSError:
        return _error_result(tool_name, "csv_read_failed")

    if len(raw) > _MAX_BYTES:
        return _error_result(tool_name, "csv_input_too_large")

    max_rows = _bounded_int(arguments.get("maxRows"), default=_DEFAULT_MAX_ROWS, maximum=_MAX_ROWS)
    max_cols = _bounded_int(arguments.get("maxCols"), default=_DEFAULT_MAX_COLS, maximum=_MAX_COLS)
    try:
        rows, source_row_count, source_col_count, redacted = _parse_csv_rows(
            raw,
            max_rows=max_rows,
            max_cols=max_cols,
        )
    except _SpreadsheetDataError as error:
        return _error_result(tool_name, error.reason_code)
    output = {
        "rows": rows,
        "rowCount": len(rows),
        "columnCount": max((len(row) for row in rows), default=0),
        "truncated": source_row_count > len(rows) or source_col_count > max_cols,
        "contentDigest": _digest(raw),
        "byteCount": len(raw),
    }
    return ToolResult(
        status="ok",
        output=output,
        llmOutput=output,
        transcriptOutput={
            "toolName": tool_name,
            "rowCount": output["rowCount"],
            "columnCount": output["columnCount"],
            "contentDigest": output["contentDigest"],
        },
        metadata={
            **_base_metadata(tool_name, permission_class="read", mutates_workspace=False),
            "contentDigest": output["contentDigest"],
            "byteCount": output["byteCount"],
            "rowCount": output["rowCount"],
            "columnCount": output["columnCount"],
            "pathRef": resolved.path_ref,
            "redactionStatus": "redacted" if redacted else "no_redaction_needed",
        },
    )


def csv_write(arguments: Mapping[str, object], context: ToolContext) -> ToolResult:
    tool_name = "csv_write"
    path_text = _string_arg(arguments, "path")
    if path_text is None:
        return _blocked_result(tool_name, "path_required")
    rows_result = _coerce_rows(arguments.get("rows"), tool_name=tool_name)
    if isinstance(rows_result, ToolResult):
        return rows_result

    try:
        root = _workspace_root(context)
        resolved = _resolve_workspace_path(root, path_text, must_exist=False)
        extension_error = _extension_error(resolved.relative)
        if extension_error is not None:
            return _blocked_result(tool_name, extension_error, _extension_message(extension_error))
    except _SpreadsheetPolicyError as error:
        return _blocked_result(tool_name, error.reason_code)

    csv_blob = _render_csv(rows_result)
    if len(csv_blob) > _MAX_BYTES:
        return _error_result(tool_name, "csv_input_too_large")

    try:
        resolved.path.parent.mkdir(parents=True, exist_ok=True)
        resolved.path.write_bytes(csv_blob)
    except OSError:
        return _error_result(tool_name, "csv_write_failed")
    content_digest = _digest(csv_blob)
    redacted_input = _rows_contain_sensitive_text(rows_result)
    output_material = {
        "contentDigest": content_digest,
        "byteCount": len(csv_blob),
        "rowCount": len(rows_result),
        "columnCount": max((len(row) for row in rows_result), default=0),
    }
    output_digest = _digest(output_material)
    artifact_ref = f"artifact:csv:{_short_digest(content_digest)}"
    receipt = {
        "kind": "local_csv_artifact",
        "artifactRef": artifact_ref,
        "contentDigest": content_digest,
        "outputDigest": output_digest,
        "byteCount": output_material["byteCount"],
        "rowCount": output_material["rowCount"],
        "columnCount": output_material["columnCount"],
        "localOnly": True,
        "deliveryClaimed": False,
        "liveAttachmentEnabled": False,
        "redactionStatus": "redacted" if redacted_input else "no_redaction_needed",
    }
    output = {
        "artifactRef": artifact_ref,
        "contentDigest": content_digest,
        "outputDigest": output_digest,
        "byteCount": output_material["byteCount"],
        "rowCount": output_material["rowCount"],
        "columnCount": output_material["columnCount"],
    }
    return ToolResult(
        status="ok",
        output=output,
        llmOutput=output,
        transcriptOutput=output,
        artifactRefs=(artifact_ref,),
        metadata={
            **_base_metadata(tool_name, permission_class="write", mutates_workspace=True),
            "localArtifactReceipt": receipt,
            "contentDigest": content_digest,
            "outputDigest": output_digest,
            "byteCount": output_material["byteCount"],
            "rowCount": output_material["rowCount"],
            "columnCount": output_material["columnCount"],
            "pathRef": resolved.path_ref,
            "redactionStatus": "redacted" if redacted_input else "no_redaction_needed",
        },
    )


def spreadsheet_preview(arguments: Mapping[str, object], context: ToolContext) -> ToolResult:
    del context
    tool_name = "spreadsheet_preview"
    rows_source = arguments.get("rows")
    if rows_source is None:
        rows_source = _rows_from_csv_read_result(arguments.get("csvReadResult"))
    rows_result = _coerce_rows(rows_source, tool_name=tool_name)
    if isinstance(rows_result, ToolResult):
        return rows_result

    sanitized_rows, redacted = _sanitize_table_rows(rows_result)
    max_rows = _bounded_int(arguments.get("maxRows"), default=10, maximum=50)
    max_cols = _bounded_int(arguments.get("maxCols"), default=8, maximum=20)
    selected_rows: list[list[str]] = []
    preview_truncated = False
    for row in sanitized_rows[:max_rows]:
        capped_row, row_truncated = _cap_preview_row(row[:max_cols])
        selected_rows.append(capped_row)
        preview_truncated = preview_truncated or row_truncated
    width = max((len(row) for row in selected_rows), default=0)
    normalized_rows = [row + [""] * (width - len(row)) for row in selected_rows]
    markdown, markdown_truncated = _cap_markdown(_markdown_table(normalized_rows))
    preview_truncated = preview_truncated or markdown_truncated
    output = {
        "markdown": markdown,
        "rowCount": len(normalized_rows),
        "columnCount": width,
        "truncated": len(rows_result) > len(normalized_rows)
        or any(len(row) > max_cols for row in rows_result)
        or preview_truncated,
    }
    return ToolResult(
        status="ok",
        output=output,
        llmOutput=output,
        transcriptOutput=output,
        metadata={
            **_base_metadata(tool_name, permission_class="meta", mutates_workspace=False),
            "rowCount": output["rowCount"],
            "columnCount": output["columnCount"],
            "redactionStatus": "redacted" if redacted else "no_redaction_needed",
            "previewTruncated": preview_truncated,
        },
    )


def _workspace_root(context: ToolContext) -> Path:
    if not context.workspace_root:
        raise _SpreadsheetPolicyError("workspace_root_required")
    root = Path(context.workspace_root).resolve()
    if not root.is_dir():
        raise _SpreadsheetPolicyError("workspace_root_required")
    return root


def _resolve_workspace_path(root: Path, path_text: str, *, must_exist: bool) -> _ResolvedPath:
    normalized = _normalize_relative(path_text)
    if not normalized:
        raise _SpreadsheetPolicyError("path_required")
    if _is_workspace_escape(path_text):
        raise _SpreadsheetPolicyError("path_escapes_workspace")
    if _is_sensitive_relative_path(normalized):
        raise _SpreadsheetPolicyError("secret_path_denied")
    candidate = root / normalized
    _reject_symlink_components(root, normalized)
    if must_exist and not candidate.exists():
        raise _SpreadsheetPolicyError("path_not_found")
    if candidate.is_symlink():
        raise _SpreadsheetPolicyError("path_symlink_denied")
    if candidate.exists():
        resolved = candidate.resolve()
        if root not in (resolved, *resolved.parents):
            raise _SpreadsheetPolicyError("path_symlink_escape_denied")
        if must_exist and not resolved.is_file():
            raise _SpreadsheetPolicyError("path_not_readable_file")
    return _ResolvedPath(
        path=candidate.resolve() if candidate.exists() else candidate,
        relative=normalized,
        path_ref=f"file:{_short_digest(normalized)}",
    )


def _reject_symlink_components(root: Path, normalized: str) -> None:
    current = root
    for part in Path(normalized).parts:
        current = current / part
        if current.is_symlink():
            raise _SpreadsheetPolicyError("path_symlink_denied")


def _normalize_relative(path_text: str) -> str:
    text = str(path_text).strip().replace("\\", "/")
    if _has_invalid_path_control(text):
        raise _SpreadsheetPolicyError("path_invalid")
    normalized = posixpath.normpath(text)
    return "" if normalized == "." else normalized


def _is_workspace_escape(path_text: str) -> bool:
    text = path_text.strip()
    if text.startswith(("/", "~")):
        return True
    if re.match(r"^[A-Za-z]:/", text.replace("\\", "/")) is not None:
        return True
    normalized = _normalize_relative(text)
    if normalized.startswith("/"):
        return True
    slash_path = text.replace("\\", "/")
    return normalized == ".." or normalized.startswith("../") or "/../" in f"/{slash_path}/"


def _has_invalid_path_control(value: str) -> bool:
    return any(ord(char) < 32 for char in value)


def _is_sensitive_relative_path(relative: str) -> bool:
    normalized = relative.replace("\\", "/").strip().lower()
    parts = [part for part in normalized.split("/") if part]
    return any(
        part in {".", ".."} or _SENSITIVE_PATH_PART_RE.search(part) is not None
        for part in parts
    )


def _extension_error(relative: str) -> str | None:
    suffix = Path(relative).suffix.casefold()
    if suffix == ".xlsx":
        return "xlsx_unsupported_dependency_approval_required"
    if suffix != ".csv":
        return "csv_extension_required"
    return None


def _extension_message(reason: str) -> str:
    if reason == "xlsx_unsupported_dependency_approval_required":
        return "xlsx support is unsupported without dependency approval"
    return "csv extension required"


def _read_bounded_bytes(path: Path) -> bytes:
    with path.open("rb") as handle:
        return handle.read(_MAX_BYTES + 1)


def _parse_csv_rows(
    raw: bytes,
    *,
    max_rows: int,
    max_cols: int,
) -> tuple[list[list[str]], int, int, bool]:
    try:
        text = raw.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError as exc:
        raise _SpreadsheetDataError("csv_decode_error") from exc
    reader = csv.reader(io.StringIO(text, newline=""), strict=True)
    raw_rows: list[list[str]] = []
    source_row_count = 0
    source_col_count = 0
    cell_count = 0
    try:
        for row in reader:
            source_row_count += 1
            source_col_count = max(source_col_count, len(row))
            if source_row_count <= max_rows:
                selected = row[:max_cols]
                cell_count += len(selected)
                if cell_count > _MAX_CELLS:
                    raise _SpreadsheetDataError("csv_input_too_large")
                raw_rows.append([str(cell) for cell in selected])
    except csv.Error as exc:
        raise _SpreadsheetDataError("csv_parse_error") from exc
    rows, redacted = _sanitize_table_rows(raw_rows)
    return rows, source_row_count, source_col_count, redacted


def _coerce_rows(value: object, *, tool_name: str) -> list[list[str]] | ToolResult:
    if not isinstance(value, Sequence) or isinstance(value, str | bytes | bytearray):
        return _error_result(tool_name, "invalid_rows_shape")
    rows: list[list[str]] = []
    cell_count = 0
    estimated_bytes = 0
    for row in value:
        if not isinstance(row, Sequence) or isinstance(row, str | bytes | bytearray):
            return _error_result(tool_name, "invalid_rows_shape")
        coerced_row: list[str] = []
        for cell in row:
            if cell is None:
                coerced = ""
            elif isinstance(cell, str | int | float | bool):
                coerced = str(cell)
            else:
                return _error_result(tool_name, "invalid_rows_shape")
            estimated_bytes += _estimated_csv_cell_bytes(coerced)
            if estimated_bytes > _MAX_BYTES:
                return _error_result(tool_name, "csv_input_too_large")
            coerced_row.append(coerced)
            cell_count += 1
            if cell_count > _MAX_CELLS:
                return _error_result(tool_name, "csv_input_too_large")
        if len(coerced_row) > _MAX_COLS:
            return _error_result(tool_name, "csv_input_too_large")
        rows.append(coerced_row)
        estimated_bytes += max(len(coerced_row) - 1, 0) + 2
        if estimated_bytes > _MAX_BYTES:
            return _error_result(tool_name, "csv_input_too_large")
        if len(rows) > _MAX_ROWS:
            return _error_result(tool_name, "csv_input_too_large")
    return rows


def _rows_from_csv_read_result(value: object) -> object:
    if not isinstance(value, Mapping):
        return value
    output = value.get("output")
    if isinstance(output, Mapping):
        return output.get("rows")
    return value.get("rows")


def _sanitize_row(row: Sequence[str]) -> tuple[list[str], bool]:
    sanitized: list[str] = []
    redacted = False
    for cell in row:
        clean, cell_redacted = _sanitize_text(str(cell))
        sanitized.append(clean)
        redacted = redacted or cell_redacted
    return sanitized, redacted


def _sanitize_table_rows(rows: Sequence[Sequence[str]]) -> tuple[list[list[str]], bool]:
    sensitive_columns: set[int] = set()
    if rows:
        for index, cell in enumerate(rows[0]):
            if _is_sensitive_table_key(str(cell)):
                sensitive_columns.add(index)

    sanitized_rows: list[list[str]] = []
    redacted = False
    for row_index, row in enumerate(rows):
        row_key_sensitive = bool(row) and _is_sensitive_table_key(str(row[0]))
        sanitized_row: list[str] = []
        for cell_index, cell in enumerate(row):
            clean, cell_redacted = _sanitize_text(str(cell))
            table_redacted = False
            if cell_index in sensitive_columns or row_key_sensitive:
                if cell_index == 0 or row_index > 0 or row_key_sensitive:
                    if clean:
                        clean = "[redacted]"
                        table_redacted = True
            sanitized_row.append(clean)
            redacted = redacted or cell_redacted or table_redacted
        sanitized_rows.append(sanitized_row)
    return sanitized_rows, redacted


def _sanitize_text(value: str) -> tuple[str, bool]:
    redacted = _SENSITIVE_TEXT_RE.sub("[redacted]", value)
    return redacted, redacted != value


def _rows_contain_sensitive_text(rows: Sequence[Sequence[str]]) -> bool:
    return _sanitize_table_rows(rows)[1]


def _is_sensitive_table_key(value: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]", "", value.casefold())
    if not normalized:
        return False
    return any(fragment in normalized for fragment in _SENSITIVE_TABLE_KEY_FRAGMENTS)


def _cap_preview_row(row: Sequence[str]) -> tuple[list[str], bool]:
    capped: list[str] = []
    truncated = False
    for cell in row:
        if len(cell) > _MAX_PREVIEW_CELL_CHARS:
            capped.append(cell[:_MAX_PREVIEW_CELL_CHARS])
            truncated = True
        else:
            capped.append(cell)
    return capped, truncated


def _cap_markdown(markdown: str) -> tuple[str, bool]:
    blob = markdown.encode("utf-8")
    if len(blob) <= _MAX_PREVIEW_MARKDOWN_BYTES:
        return markdown, False
    limited = blob[:_MAX_PREVIEW_MARKDOWN_BYTES].decode("utf-8", errors="ignore")
    return limited, True


def _estimated_csv_cell_bytes(value: str) -> int:
    extra_quotes = value.count('"')
    needs_quotes = any(char in value for char in (",", '"', "\r", "\n"))
    wrapper = 2 if needs_quotes else 0
    return len(value.encode("utf-8")) + extra_quotes + wrapper


def _render_csv(rows: list[list[str]]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def _markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    header = rows[0]
    divider = ["---"] * len(header)
    lines = [_markdown_row(header), _markdown_row(divider)]
    lines.extend(_markdown_row(row) for row in rows[1:])
    return "\n".join(lines)


def _markdown_row(row: Sequence[str]) -> str:
    escaped = [cell.replace("|", "\\|").replace("\n", " ") for cell in row]
    return "| " + " | ".join(escaped) + " |"


def _string_arg(arguments: Mapping[str, object], name: str) -> str | None:
    value = arguments.get(name)
    if isinstance(value, str):
        return value
    return None


def _bounded_int(value: object, *, default: int, maximum: int) -> int:
    if isinstance(value, bool):
        return default
    if isinstance(value, int):
        return min(max(value, 1), maximum)
    if isinstance(value, str) and value.isdecimal():
        return min(max(int(value), 1), maximum)
    return default


def _base_metadata(
    tool_name: str,
    *,
    permission_class: str,
    mutates_workspace: bool,
) -> dict[str, object]:
    return {
        "toolName": tool_name,
        "permissionClass": permission_class,
        "dangerous": False,
        "mutatesWorkspace": mutates_workspace,
        "localOnly": True,
        "subprocessFree": True,
        "networkAllowed": False,
        "deliveryClaimed": False,
        "liveAttachmentEnabled": False,
    }


def _blocked_result(tool_name: str, reason: str, message: str | None = None) -> ToolResult:
    return ToolResult(
        status="blocked",
        errorCode=reason,
        errorMessage=message or reason.replace("_", " "),
        metadata={
            **_base_metadata(
                tool_name,
                permission_class=_permission_class_for(tool_name),
                mutates_workspace=False,
            ),
            "reason": reason,
        },
    )


def _error_result(tool_name: str, reason: str) -> ToolResult:
    return ToolResult(
        status="error",
        errorCode=reason,
        errorMessage=reason.replace("_", " "),
        metadata={
            **_base_metadata(
                tool_name,
                permission_class=_permission_class_for(tool_name),
                mutates_workspace=False,
            ),
            "reason": reason,
        },
    )


def _permission_class_for(tool_name: str) -> str:
    if tool_name == "csv_read":
        return "read"
    if tool_name == "spreadsheet_preview":
        return "meta"
    return "write"


def _digest(value: object) -> str:
    if isinstance(value, bytes):
        encoded = value
    elif isinstance(value, str):
        encoded = value.encode("utf-8")
    else:
        encoded = json.dumps(
            value,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=True,
            default=repr,
        ).encode("utf-8")
    return f"sha256:{hashlib.sha256(encoded).hexdigest()}"


def _short_digest(value: object) -> str:
    return _digest(value).removeprefix("sha256:")[:_DIGEST_PREFIX_LENGTH]


def xlsx_read(arguments: Mapping[str, object], context: ToolContext) -> ToolResult:
    """Read an XLSX workbook from the workspace, returning structured rows.

    Requires the ``openpyxl`` package (``uv sync --extra files``).  When the
    package is not installed the handler returns ``status="blocked"`` so the
    manifest can safely exist in all environments.
    """
    tool_name = "xlsx_read"
    path_text = _string_arg(arguments, "path")
    if path_text is None:
        return _blocked_result(tool_name, "path_required")

    try:
        root = _workspace_root(context)
        resolved = _resolve_workspace_path(root, path_text, must_exist=True)
    except _SpreadsheetPolicyError as error:
        return _blocked_result(tool_name, error.reason_code)
    except OSError:
        return _error_result(tool_name, "xlsx_read_failed")

    if Path(resolved.relative).suffix.casefold() != ".xlsx":
        return _blocked_result(tool_name, "xlsx_extension_required")

    try:
        raw_size = resolved.path.stat().st_size
    except OSError:
        return _error_result(tool_name, "xlsx_read_failed")
    if raw_size > _MAX_BYTES:
        return _error_result(tool_name, "xlsx_input_too_large")

    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        return _blocked_result(tool_name, "xlsx_dependency_not_installed")

    try:
        wb = openpyxl.load_workbook(resolved.path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return _error_result(tool_name, "xlsx_read_failed")

    sheet_name = _string_arg(arguments, "sheetName")
    try:
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                return _blocked_result(tool_name, "xlsx_sheet_not_found")
            ws = wb[sheet_name]
        else:
            ws = wb.active
    except Exception:  # noqa: BLE001
        return _error_result(tool_name, "xlsx_read_failed")

    max_rows = _bounded_int(arguments.get("maxRows"), default=_DEFAULT_MAX_ROWS, maximum=_MAX_ROWS)
    max_cols = _bounded_int(arguments.get("maxCols"), default=_DEFAULT_MAX_COLS, maximum=_MAX_COLS)

    # Optional cell range filter: "A1:C5" restricts to min/max row+col
    cell_range_str = _string_arg(arguments, "cellRange")
    row_min: int | None = None
    row_max: int | None = None
    col_min: int | None = None
    col_max: int | None = None
    if cell_range_str is not None:
        range_result = _parse_cell_range(cell_range_str)
        if range_result is None:
            wb.close()
            return _blocked_result(tool_name, "xlsx_invalid_cell_range")
        row_min, col_min, row_max, col_max = range_result

    raw_rows: list[list[str]] = []
    source_row_count = 0
    cell_count = 0
    try:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            source_row_count += 1
            # Apply row range filter
            if row_min is not None and row_idx < row_min:
                continue
            if row_max is not None and row_idx > row_max:
                break
            if source_row_count > max_rows and row_min is None:
                continue
            # Apply column range filter
            row_list = list(row)
            if col_min is not None:
                row_list = row_list[col_min - 1 : (col_max or max_cols)]
            else:
                row_list = row_list[:max_cols]
            coerced: list[str] = []
            for cell_value in row_list:
                if cell_value is None:
                    coerced.append("")
                elif isinstance(cell_value, bool):
                    coerced.append(str(cell_value))
                elif isinstance(cell_value, int | float):
                    coerced.append(str(cell_value))
                else:
                    coerced.append(str(cell_value))
            cell_count += len(coerced)
            if cell_count > _MAX_CELLS:
                break
            raw_rows.append(coerced)
    except Exception:  # noqa: BLE001
        return _error_result(tool_name, "xlsx_read_failed")
    finally:
        wb.close()

    rows, redacted = _sanitize_table_rows(raw_rows)
    truncated = source_row_count > max_rows or any(
        len(list(row)) > max_cols for row in rows
    )
    content_digest = _digest(resolved.path.read_bytes() if raw_size <= _MAX_BYTES else b"")

    output: dict[str, object] = {
        "rows": rows,
        "rowCount": len(rows),
        "columnCount": max((len(row) for row in rows), default=0),
        "truncated": truncated,
        "contentDigest": content_digest,
        "byteCount": raw_size,
    }
    return ToolResult(
        status="ok",
        output=output,
        llmOutput=output,
        transcriptOutput={
            "toolName": tool_name,
            "rowCount": output["rowCount"],
            "columnCount": output["columnCount"],
            "contentDigest": output["contentDigest"],
        },
        metadata={
            **_base_metadata(tool_name, permission_class="read", mutates_workspace=False),
            "contentDigest": output["contentDigest"],
            "byteCount": raw_size,
            "rowCount": output["rowCount"],
            "columnCount": output["columnCount"],
            "pathRef": resolved.path_ref,
            "redactionStatus": "redacted" if redacted else "no_redaction_needed",
        },
    )


def xls_read(arguments: Mapping[str, object], context: ToolContext) -> ToolResult:
    """Read an XLS (BIFF) workbook from the workspace, returning structured rows.

    Requires the ``xlrd`` package (``uv sync --extra files``). When the
    package is not installed the handler returns ``status="blocked"`` so the
    manifest can safely exist in all environments.

    Returns the same output shape as ``xlsx_read``.
    """
    tool_name = "xls_read"
    path_text = _string_arg(arguments, "path")
    if path_text is None:
        return _blocked_result(tool_name, "path_required")

    try:
        root = _workspace_root(context)
        resolved = _resolve_workspace_path(root, path_text, must_exist=True)
    except _SpreadsheetPolicyError as error:
        return _blocked_result(tool_name, error.reason_code)
    except OSError:
        return _error_result(tool_name, "xls_read_failed")

    if Path(resolved.relative).suffix.casefold() != ".xls":
        return _blocked_result(tool_name, "xls_extension_required")

    try:
        raw_size = resolved.path.stat().st_size
    except OSError:
        return _error_result(tool_name, "xls_read_failed")
    if raw_size > _MAX_BYTES:
        return _error_result(tool_name, "xls_input_too_large")

    try:
        import xlrd  # noqa: PLC0415
    except ImportError:
        return _blocked_result(tool_name, "xls_dependency_not_installed")

    try:
        wb = xlrd.open_workbook(str(resolved.path))
    except Exception:  # noqa: BLE001
        return _error_result(tool_name, "xls_read_failed")

    sheet_name = _string_arg(arguments, "sheetName")
    try:
        if sheet_name is not None:
            ws = wb.sheet_by_name(sheet_name)
        else:
            ws = wb.sheet_by_index(0)
    except Exception:  # noqa: BLE001
        return _error_result(tool_name, "xls_read_failed")

    max_rows = _bounded_int(arguments.get("maxRows"), default=_DEFAULT_MAX_ROWS, maximum=_MAX_ROWS)
    max_cols = _bounded_int(arguments.get("maxCols"), default=_DEFAULT_MAX_COLS, maximum=_MAX_COLS)

    raw_rows: list[list[str]] = []
    source_row_count = ws.nrows
    cell_count = 0
    try:
        for row_idx in range(min(ws.nrows, max_rows)):
            row_cells = ws.row(row_idx)[:max_cols]
            coerced: list[str] = []
            for cell in row_cells:
                ctype = cell.ctype
                val = cell.value
                if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    coerced.append("")
                elif ctype == xlrd.XL_CELL_BOOLEAN:
                    coerced.append(str(bool(val)))
                elif ctype == xlrd.XL_CELL_NUMBER:
                    if isinstance(val, float) and val == int(val):
                        coerced.append(str(int(val)))
                    else:
                        coerced.append(str(val))
                elif ctype == xlrd.XL_CELL_ERROR:
                    coerced.append("")
                else:
                    coerced.append(str(val))
            cell_count += len(coerced)
            if cell_count > _MAX_CELLS:
                break
            raw_rows.append(coerced)
    except Exception:  # noqa: BLE001
        return _error_result(tool_name, "xls_read_failed")
    finally:
        try:
            wb.release_resources()
        except Exception:  # noqa: BLE001
            pass

    rows, redacted = _sanitize_table_rows(raw_rows)
    truncated = source_row_count > max_rows or any(
        ws.ncols > max_cols for _ in (None,)
    )
    content_digest = _digest(resolved.path.read_bytes() if raw_size <= _MAX_BYTES else b"")

    output: dict[str, object] = {
        "rows": rows,
        "rowCount": len(rows),
        "columnCount": max((len(row) for row in rows), default=0),
        "truncated": truncated,
        "contentDigest": content_digest,
        "byteCount": raw_size,
    }
    return ToolResult(
        status="ok",
        output=output,
        llmOutput=output,
        transcriptOutput={
            "toolName": tool_name,
            "rowCount": output["rowCount"],
            "columnCount": output["columnCount"],
            "contentDigest": output["contentDigest"],
        },
        metadata={
            **_base_metadata(tool_name, permission_class="read", mutates_workspace=False),
            "contentDigest": output["contentDigest"],
            "byteCount": raw_size,
            "rowCount": output["rowCount"],
            "columnCount": output["columnCount"],
            "pathRef": resolved.path_ref,
            "redactionStatus": "redacted" if redacted else "no_redaction_needed",
        },
    )


def xlsx_info(arguments: Mapping[str, object], context: ToolContext) -> ToolResult:
    """Return structural metadata about an XLSX workbook without reading row data.

    Lists all sheets with their row count, column count, and a preview of the
    first row (header).

    Requires the ``openpyxl`` package (``uv sync --extra files``).  When the
    package is not installed the handler returns ``status="blocked"``.

    Output fields
    -------------
    - ``sheets``: list of ``{name, rowCount, columnCount, headerPreview}``
    - ``sheetCount``: total number of sheets
    - ``contentDigest``: sha256 of the file bytes
    """
    tool_name = "xlsx_info"
    path_text = _string_arg(arguments, "path")
    if path_text is None:
        return _blocked_result(tool_name, "path_required")

    try:
        root = _workspace_root(context)
        resolved = _resolve_workspace_path(root, path_text, must_exist=True)
    except _SpreadsheetPolicyError as error:
        return _blocked_result(tool_name, error.reason_code)
    except OSError:
        return _error_result(tool_name, "xlsx_read_failed")

    if Path(resolved.relative).suffix.casefold() != ".xlsx":
        return _blocked_result(tool_name, "xlsx_extension_required")

    try:
        raw_size = resolved.path.stat().st_size
    except OSError:
        return _error_result(tool_name, "xlsx_read_failed")

    if raw_size > _MAX_BYTES:
        return _error_result(tool_name, "xlsx_input_too_large")

    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        return _blocked_result(tool_name, "xlsx_dependency_not_installed")

    try:
        wb = openpyxl.load_workbook(resolved.path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return _error_result(tool_name, "xlsx_read_failed")

    sheets_info: list[dict[str, object]] = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            row_count = 0
            col_count = 0
            header_preview: list[str] = []
            try:
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    row_list = [c for c in row]
                    row_count += 1
                    col_count = max(col_count, len(row_list))
                    if row_idx == 1:
                        header_preview = [
                            str(c) if c is not None else ""
                            for c in row_list[:_DEFAULT_MAX_COLS]
                        ]
                    if row_count >= _MAX_ROWS:
                        break
            except Exception:  # noqa: BLE001
                pass
            sheets_info.append(
                {
                    "name": name,
                    "rowCount": row_count,
                    "columnCount": col_count,
                    "headerPreview": header_preview,
                }
            )
    finally:
        wb.close()

    content_digest = _digest(resolved.path.read_bytes() if raw_size <= _MAX_BYTES else b"")
    output: dict[str, object] = {
        "sheets": sheets_info,
        "sheetCount": len(sheets_info),
        "contentDigest": content_digest,
    }
    return ToolResult(
        status="ok",
        output=output,
        llmOutput=output,
        transcriptOutput={
            "toolName": tool_name,
            "sheetCount": len(sheets_info),
            "contentDigest": content_digest,
        },
        metadata={
            **_base_metadata(tool_name, permission_class="read", mutates_workspace=False),
            "contentDigest": content_digest,
            "byteCount": raw_size,
            "sheetCount": len(sheets_info),
            "pathRef": resolved.path_ref,
        },
    )


def _parse_cell_range(
    cell_range_str: str,
) -> tuple[int, int, int, int] | None:
    """Parse an Excel-style cell range like ``'A1:C5'``.

    Returns ``(row_min, col_min, row_max, col_max)`` as 1-based integers, or
    ``None`` when the string cannot be parsed.
    """
    import re  # noqa: PLC0415

    pattern = re.compile(
        r"^([A-Za-z]{1,3})(\d+):([A-Za-z]{1,3})(\d+)$",
        re.IGNORECASE,
    )
    m = pattern.match(cell_range_str.strip())
    if not m:
        return None
    col_min = _col_letter_to_index(m.group(1))
    row_min = int(m.group(2))
    col_max = _col_letter_to_index(m.group(3))
    row_max = int(m.group(4))
    if col_min < 1 or col_max < col_min or row_min < 1 or row_max < row_min:
        return None
    return row_min, col_min, row_max, col_max


def _col_letter_to_index(letters: str) -> int:
    """Convert Excel column letters to a 1-based column index.

    E.g. ``'A'`` → 1, ``'Z'`` → 26, ``'AA'`` → 27.
    """
    result = 0
    for char in letters.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


__all__ = ["csv_read", "csv_write", "spreadsheet_preview", "xlsx_info", "xlsx_read"]
